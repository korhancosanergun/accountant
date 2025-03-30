#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UK Muhasebe Yazılımı - Veri İçe/Dışa Aktarım Modülü
Excel, CSV ve diğer formatlar için içe ve dışa aktarım işlemleri
"""

import os
import csv
import json
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
from pathlib import Path

# Modül için logger
logger = logging.getLogger(__name__)


class ImportExport:
    """Veri içe ve dışa aktarım sınıfı"""
    
    def __init__(self, ledger):
        """Sınıf başlatıcı
        
        Args:
            ledger: Muhasebe defteri nesnesi
        """
        self.ledger = ledger
    
    def export_to_excel(self, data_type, file_path, date_range=None):
        """Verileri Excel formatında dışa aktar
        
        Args:
            data_type: Veri tipi ("transactions", "invoices", "chart_of_accounts", "income_expenses", "vat_returns")
            file_path: Kaydedilecek dosya yolu
            date_range: Tarih aralığı (start_date, end_date) veya None
            
        Returns:
            bool: İşlem başarılı mı
        """
        try:
            # Veri tipine göre dışa aktarılacak verileri al
            if data_type == "transactions":
                data = self._prepare_transactions_data(date_range)
                sheet_name = "Muhasebe Defteri"
            elif data_type == "invoices":
                data = self._prepare_invoices_data(date_range)
                sheet_name = "Faturalar"
            elif data_type == "chart_of_accounts":
                data = self._prepare_chart_of_accounts_data()
                sheet_name = "Hesap Planı"
            elif data_type == "income_expenses":
                data = self._prepare_income_expenses_data(date_range)
                sheet_name = "Gelir Gider"
            elif data_type == "vat_returns":
                data = self._prepare_vat_returns_data(date_range)
                sheet_name = "KDV Beyanları"
            else:
                raise ValueError(f"Geçersiz veri tipi: {data_type}")
            
            # Boş Excel çalışma kitabı oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Stil tanımları
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            thin_border = Side(border_style="thin", color="000000")
            border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            
            # Başlıkları yaz
            for col_idx, header in enumerate(data["headers"], 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Verileri yaz
            for row_idx, row_data in enumerate(data["rows"], 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.border = border
                    
                    # Hücre tipine göre hizalama
                    if isinstance(cell_value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                    elif isinstance(cell_value, str) and cell_value.strip().startswith("£"):
                        cell.alignment = Alignment(horizontal="right")
            
            # Sütun genişliklerini içeriğe göre ayarla
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = max(max_length + 2, 10)  # Minimum 10 karakterlik genişlik
                ws.column_dimensions[column].width = adjusted_width
            
            # Excel dosyasını kaydet
            wb.save(file_path)
            
            logger.info(f"{data_type} verileri Excel formatında dışa aktarıldı: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel dışa aktarımında hata: {e}")
            return False
    
    def export_to_csv(self, data_type, file_path, date_range=None):
        """Verileri CSV formatında dışa aktar
        
        Args:
            data_type: Veri tipi ("transactions", "invoices", "chart_of_accounts", "income_expenses", "vat_returns")
            file_path: Kaydedilecek dosya yolu
            date_range: Tarih aralığı (start_date, end_date) veya None
            
        Returns:
            bool: İşlem başarılı mı
        """
        try:
            # Veri tipine göre dışa aktarılacak verileri al
            if data_type == "transactions":
                data = self._prepare_transactions_data(date_range)
            elif data_type == "invoices":
                data = self._prepare_invoices_data(date_range)
            elif data_type == "chart_of_accounts":
                data = self._prepare_chart_of_accounts_data()
            elif data_type == "income_expenses":
                data = self._prepare_income_expenses_data(date_range)
            elif data_type == "vat_returns":
                data = self._prepare_vat_returns_data(date_range)
            else:
                raise ValueError(f"Geçersiz veri tipi: {data_type}")
            
            # CSV dosyasını yaz
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Başlıkları yaz
                writer.writerow(data["headers"])
                
                # Verileri yaz
                writer.writerows(data["rows"])
            
            logger.info(f"{data_type} verileri CSV formatında dışa aktarıldı: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"CSV dışa aktarımında hata: {e}")
            return False
    
    def export_to_json(self, data_type, file_path, date_range=None):
        """Verileri JSON formatında dışa aktar
        
        Args:
            data_type: Veri tipi ("transactions", "invoices", "chart_of_accounts", "income_expenses", "vat_returns")
            file_path: Kaydedilecek dosya yolu
            date_range: Tarih aralığı (start_date, end_date) veya None
            
        Returns:
            bool: İşlem başarılı mı
        """
        try:
            # Veri tipine göre dışa aktarılacak verileri al
            if data_type == "transactions":
                data = self.ledger.get_all_transactions()
                if date_range:
                    # Tarih filtreleme
                    start_date, end_date = date_range
                    data = [t for t in data if start_date <= t.get("date", "") <= end_date]
            elif data_type == "invoices":
                data = self.ledger.get_all_invoices()
                if date_range:
                    # Tarih filtreleme
                    start_date, end_date = date_range
                    data = [i for i in data if start_date <= i.get("date", "") <= end_date]
            elif data_type == "chart_of_accounts":
                data = self.ledger.get_chart_of_accounts()
            elif data_type == "income_expenses":
                data = self.ledger.get_income_expenses()
                if date_range:
                    # Tarih filtreleme
                    start_date, end_date = date_range
                    data = [ie for ie in data if start_date <= ie.get("date", "") <= end_date]
            elif data_type == "vat_returns":
                data = self.ledger.get_vat_returns()
                if date_range:
                    # Tarih filtreleme
                    start_date, end_date = date_range
                    data = [v for v in data if start_date <= v.get("period_start", "") <= end_date]
            else:
                raise ValueError(f"Geçersiz veri tipi: {data_type}")
            
            # JSON dosyasını yaz
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"{data_type} verileri JSON formatında dışa aktarıldı: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"JSON dışa aktarımında hata: {e}")
            return False
    
    def import_from_excel(self, data_type, file_path):
        """Excel dosyasından veri içe aktar
        
        Args:
            data_type: Veri tipi ("transactions", "invoices", "chart_of_accounts")
            file_path: İçe aktarılacak dosya yolu
            
        Returns:
            tuple: (başarılı mı, içe aktarılan kayıt sayısı, hata mesajı)
        """
        try:
            # Excel dosyasını yükle
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Başlıkları al
            headers = [cell.value for cell in ws[1]]
            
            # Verileri al
            rows = []
            for row in ws.iter_rows(min_row=2):
                row_data = [cell.value for cell in row]
                rows.append(row_data)
            
            # Veri tipine göre içe aktarım
            if data_type == "transactions":
                return self._import_transactions(headers, rows)
            elif data_type == "invoices":
                return self._import_invoices(headers, rows)
            elif data_type == "chart_of_accounts":
                return self._import_chart_of_accounts(headers, rows)
            else:
                raise ValueError(f"Geçersiz veri tipi: {data_type}")
            
        except Exception as e:
            logger.error(f"Excel içe aktarımında hata: {e}")
            return False, 0, str(e)
    
    def import_from_csv(self, data_type, file_path):
        """CSV dosyasından veri içe aktar
        
        Args:
            data_type: Veri tipi ("transactions", "invoices", "chart_of_accounts")
            file_path: İçe aktarılacak dosya yolu
            
        Returns:
            tuple: (başarılı mı, içe aktarılan kayıt sayısı, hata mesajı)
        """
        try:
            # CSV dosyasını yükle
            with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                
                # Başlıkları al
                headers = next(reader)
                
                # Verileri al
                rows = list(reader)
            
            # Veri tipine göre içe aktarım
            if data_type == "transactions":
                return self._import_transactions(headers, rows)
            elif data_type == "invoices":
                return self._import_invoices(headers, rows)
            elif data_type == "chart_of_accounts":
                return self._import_chart_of_accounts(headers, rows)
            else:
                raise ValueError(f"Geçersiz veri tipi: {data_type}")
            
        except Exception as e:
            logger.error(f"CSV içe aktarımında hata: {e}")
            return False, 0, str(e)
    
    def _prepare_transactions_data(self, date_range=None):
        """İşlem verilerini dışa aktarım için hazırla
        
        Args:
            date_range: Tarih aralığı (start_date, end_date) veya None
            
        Returns:
            dict: Başlıklar ve satır verileri içeren sözlük
        """
        # Verileri al
        transactions = self.ledger.get_all_transactions()
        
        # Tarih filtresi uygula
        if date_range:
            start_date, end_date = date_range
            transactions = [t for t in transactions if start_date <= t.get("date", "") <= end_date]
        
        # Tarihe göre sırala
        transactions.sort(key=lambda x: x.get("date", ""), reverse=True)
        
        # Başlıklar
        headers = [
            "Tarih", "Belge No", "Açıklama", "Hesap Kodu", "Hesap Adı", 
            "Borç", "Alacak", "KDV", "Durum", "İşlem Tipi", "Notlar"
        ]
        
        # Satırlar
        rows = []
        for trans in transactions:
            # Hesap adını al
            account_code = trans.get("account", "")
            account = self.ledger.get_account_by_code(account_code)
            account_name = account.get("name", "") if account else ""
            
            # Satır verisi
            row = [
                trans.get("date", ""),
                trans.get("document_number", ""),
                trans.get("description", ""),
                account_code,
                account_name,
                trans.get("debit", 0),
                trans.get("credit", 0),
                trans.get("vat", 0),
                trans.get("status", ""),
                trans.get("transaction_type", ""),
                trans.get("notes", "")
            ]
            rows.append(row)
        
        return {"headers": headers, "rows": rows}
    
    def _prepare_invoices_data(self, date_range=None):
        """Fatura verilerini dışa aktarım için hazırla
        
        Args:
            date_range: Tarih aralığı (start_date, end_date) veya None
            
        Returns:
            dict: Başlıklar ve satır verileri içeren sözlük
        """
        # Verileri al
        invoices = self.ledger.get_all_invoices()
        
        # Tarih filtresi uygula
        if date_range:
            start_date, end_date = date_range
            invoices = [i for i in invoices if start_date <= i.get("date", "") <= end_date]
        
        # Tarihe göre sırala
        invoices.sort(key=lambda x: x.get("date", ""), reverse=True)
        
        # Başlıklar
        headers = [
            "Fatura No", "Tarih", "Vade Tarihi", "Müşteri/Tedarikçi",
            "Tutar", "KDV", "Toplam", "Durum", "Ödeme Tarihi", "Tip"
        ]
        
        # Satırlar
        rows = []
        for inv in invoices:
            # Satır verisi
            row = [
                inv.get("invoice_number", ""),
                inv.get("date", ""),
                inv.get("due_date", ""),
                inv.get("entity_name", ""),
                inv.get("amount", 0),
                inv.get("vat", 0),
                inv.get("amount", 0) + inv.get("vat", 0),
                inv.get("payment_status", ""),
                inv.get("payment_date", ""),
                "Satış" if inv.get("type") == "sales" else "Alış"
            ]
            rows.append(row)
        
        return {"headers": headers, "rows": rows}
    
    def _prepare_chart_of_accounts_data(self):
        """Hesap planı verilerini dışa aktarım için hazırla
        
        Returns:
            dict: Başlıklar ve satır verileri içeren sözlük
        """
        # Verileri al
        accounts = self.ledger.get_chart_of_accounts()
        
        # Kod sırasına göre sırala
        accounts.sort(key=lambda x: x.get("code", ""))
        
        # Başlıklar
        headers = [
            "Hesap Kodu", "Hesap Adı", "Tür", "Kategori", "KDV Oranı", "Bakiye"
        ]
        
        # Satırlar
        rows = []
        for acc in accounts:
            # Hesap türünü çevir
            account_type_map = {
                "asset": "Varlık",
                "liability": "Borç",
                "equity": "Öz Sermaye",
                "income": "Gelir",
                "expense": "Gider"
            }
            account_type = account_type_map.get(acc.get("type", ""), acc.get("type", ""))
            
            # Kategoriyi çevir
            category_map = {
                "current_asset": "Dönen Varlık",
                "fixed_asset": "Duran Varlık",
                "current_liability": "Kısa Vadeli Borç",
                "long_term_liability": "Uzun Vadeli Borç",
                "equity": "Öz Sermaye",
                "revenue": "Hasılat",
                "cost_of_sales": "Satış Maliyeti",
                "operating_expense": "Faaliyet Gideri",
                "financial_expense": "Finansman Gideri",
                "other_income": "Diğer Gelir"
            }
            category = category_map.get(acc.get("category", ""), acc.get("category", ""))
            
            # Satır verisi
            row = [
                acc.get("code", ""),
                acc.get("name", ""),
                account_type,
                category,
                f"{acc.get('vat_rate', 0)}%",
                acc.get("balance", 0)
            ]
            rows.append(row)
        
        return {"headers": headers, "rows": rows}
    
    def _prepare_income_expenses_data(self, date_range=None):
        """Gelir/Gider verilerini dışa aktarım için hazırla
        
        Args:
            date_range: Tarih aralığı (start_date, end_date) veya None
            
        Returns:
            dict: Başlıklar ve satır verileri içeren sözlük
        """
        # Verileri al
        items = self.ledger.get_income_expenses()
        
        # Tarih filtresi uygula
        if date_range:
            start_date, end_date = date_range
            items = [i for i in items if start_date <= i.get("date", "") <= end_date]
        
        # Tarihe göre sırala
        items.sort(key=lambda x: x.get("date", ""), reverse=True)
        
        # Başlıklar
        headers = [
            "Tarih", "Kategori", "Açıklama", "Gelir", "Gider", "KDV",
            "Toplam", "Ödeme Yöntemi", "Fiş/Fatura No", "Durum"
        ]
        
        # Satırlar
        rows = []
        for item in items:
            # Toplam hesapla
            income = item.get("income", 0)
            expense = item.get("expense", 0)
            vat = item.get("vat", 0)
            total = income - expense
            
            # Kategoriyi çevir
            category_map = {
                "office": "Ofis Giderleri",
                "travel": "Seyahat ve Konaklama",
                "marketing": "Pazarlama ve Reklam",
                "rent": "Kira",
                "utilities": "Faturalar",
                "software": "Yazılım ve Abonelikler",
                "professional": "Profesyonel Hizmetler",
                "salary": "Maaşlar",
                "bank": "Banka ve Finansman Giderleri",
                "sales": "Satışlar",
                "other": "Diğer"
            }
            category = category_map.get(item.get("category", ""), item.get("category", ""))
            
            # Ödeme yöntemini çevir
            payment_method_map = {
                "cash": "Nakit",
                "bank": "Banka",
                "credit_card": "Kredi Kartı",
                "other": "Diğer"
            }
            payment_method = payment_method_map.get(item.get("payment_method", ""), item.get("payment_method", ""))
            
            # Durumu çevir
            status_map = {
                "paid": "Ödenmiş",
                "unpaid": "Ödenmemiş"
            }
            status = status_map.get(item.get("status", ""), item.get("status", ""))
            
            # Satır verisi
            row = [
                item.get("date", ""),
                category,
                item.get("description", ""),
                income,
                expense,
                vat,
                total,
                payment_method,
                item.get("receipt_number", ""),
                status
            ]
            rows.append(row)
        
        return {"headers": headers, "rows": rows}
    
    def _prepare_vat_returns_data(self, date_range=None):
        """KDV beyanı verilerini dışa aktarım için hazırla
        
        Args:
            date_range: Tarih aralığı (start_date, end_date) veya None
            
        Returns:
            dict: Başlıklar ve satır verileri içeren sözlük
        """
        # Verileri al
        vat_returns = self.ledger.get_vat_returns()
        
        # Tarih filtresi uygula
        if date_range:
            start_date, end_date = date_range
            vat_returns = [v for v in vat_returns if start_date <= v.get("period_start", "") <= end_date]
        
        # Dönem başlangıç tarihine göre sırala
        vat_returns.sort(key=lambda x: x.get("period_start", ""), reverse=True)
        
        # Başlıklar
        headers = [
            "Dönem Başlangıç", "Dönem Bitiş", "Gönderim Tarihi", "Satışlardan KDV",
            "AB Alımlarından KDV", "Toplam KDV", "İndirilecek KDV", "Net KDV",
            "Toplam Satışlar", "Toplam Alımlar", "Durum"
        ]
        
        # Satırlar
        rows = []
        for vat in vat_returns:
            # Gönderim tarihini formatla
            submission_date = vat.get("submission_date", "")
            try:
                if submission_date:
                    dt = datetime.fromisoformat(submission_date)
                    submission_date = dt.strftime("%Y-%m-%d %H:%M")
            except ValueError:
                pass
            
            # Satır verisi
            row = [
                vat.get("period_start", ""),
                vat.get("period_end", ""),
                submission_date,
                vat.get("vat_due_sales", 0),
                vat.get("vat_due_acquisitions", 0),
                vat.get("total_vat_due", 0),
                vat.get("vat_reclaimed", 0),
                vat.get("net_vat_due", 0),
                vat.get("total_sales_ex_vat", 0),
                vat.get("total_purchases_ex_vat", 0),
                vat.get("status", "")
            ]
            rows.append(row)
        
        return {"headers": headers, "rows": rows}
    
    def _import_transactions(self, headers, rows):
        """İşlem verilerini içe aktar
        
        Args:
            headers: Başlık satırı
            rows: Veri satırları
            
        Returns:
            tuple: (başarılı mı, içe aktarılan kayıt sayısı, hata mesajı)
        """
        try:
            # Gerekli sütunları doğrula
            required_columns = ["Tarih", "Açıklama", "Hesap Kodu", "Borç", "Alacak"]
            for col in required_columns:
                if col not in headers:
                    return False, 0, f"Gerekli sütun bulunamadı: {col}"
            
            # Sütun indekslerini bul
            date_idx = headers.index("Tarih")
            doc_idx = headers.index("Belge No") if "Belge No" in headers else None
            desc_idx = headers.index("Açıklama")
            acc_idx = headers.index("Hesap Kodu")
            debit_idx = headers.index("Borç")
            credit_idx = headers.index("Alacak")
            vat_idx = headers.index("KDV") if "KDV" in headers else None
            status_idx = headers.index("Durum") if "Durum" in headers else None
            type_idx = headers.index("İşlem Tipi") if "İşlem Tipi" in headers else None
            notes_idx = headers.index("Notlar") if "Notlar" in headers else None
            
            # Başarılı içe aktarım sayacı
            success_count = 0
            
            # Her satır için
            for row in rows:
                try:
                    # Boş satırları atla
                    if not row or len(row) < max(date_idx, desc_idx, acc_idx, debit_idx, credit_idx) + 1:
                        continue
                    
                    # İşlem verilerini oluştur
                    transaction = {
                        "date": row[date_idx],
                        "description": row[desc_idx],
                        "account": row[acc_idx],
                        "debit": float(row[debit_idx] or 0),
                        "credit": float(row[credit_idx] or 0)
                    }
                    
                    # Opsiyonel alanlar
                    if doc_idx is not None and doc_idx < len(row):
                        transaction["document_number"] = row[doc_idx]
                    
                    if vat_idx is not None and vat_idx < len(row):
                        transaction["vat"] = float(row[vat_idx] or 0)
                    
                    if status_idx is not None and status_idx < len(row):
                        transaction["status"] = row[status_idx]
                    
                    if type_idx is not None and type_idx < len(row):
                        transaction["transaction_type"] = row[type_idx]
                    
                    if notes_idx is not None and notes_idx < len(row):
                        transaction["notes"] = row[notes_idx]
                    
                    # İşlemi ekle
                    self.ledger.add_transaction(transaction)
                    success_count += 1
                    
                except Exception as e:
                    logger.warning(f"Satır içe aktarımında hata: {e}")
                    continue
            
            return True, success_count, ""
            
        except Exception as e:
            logger.error(f"İşlem içe aktarımında hata: {e}")
            return False, 0, str(e)
    
    def _import_invoices(self, headers, rows):
        """Fatura verilerini içe aktar
        
        Args:
            headers: Başlık satırı
            rows: Veri satırları
            
        Returns:
            tuple: (başarılı mı, içe aktarılan kayıt sayısı, hata mesajı)
        """
        try:
            # Gerekli sütunları doğrula
            required_columns = ["Fatura No", "Tarih", "Müşteri/Tedarikçi", "Tutar", "Tip"]
            for col in required_columns:
                if col not in headers:
                    return False, 0, f"Gerekli sütun bulunamadı: {col}"
            
            # Sütun indekslerini bul
            inv_idx = headers.index("Fatura No")
            date_idx = headers.index("Tarih")
            due_idx = headers.index("Vade Tarihi") if "Vade Tarihi" in headers else None
            entity_idx = headers.index("Müşteri/Tedarikçi")
            amount_idx = headers.index("Tutar")
            vat_idx = headers.index("KDV") if "KDV" in headers else None
            status_idx = headers.index("Durum") if "Durum" in headers else None
            type_idx = headers.index("Tip")
            
            # Başarılı içe aktarım sayacı
            success_count = 0
            
            # Her satır için
            for row in rows:
                try:
                    # Boş satırları atla
                    if not row or len(row) < max(inv_idx, date_idx, entity_idx, amount_idx, type_idx) + 1:
                        continue
                    
                    # Fatura tipini belirle
                    invoice_type = "sales" if row[type_idx] == "Satış" else "purchase"
                    
                    # Fatura verilerini oluştur
                    invoice = {
                        "invoice_number": row[inv_idx],
                        "date": row[date_idx],
                        "entity_name": row[entity_idx],
                        "amount": float(row[amount_idx] or 0),
                        "type": invoice_type
                    }
                    
                    # Opsiyonel alanlar
                    if due_idx is not None and due_idx < len(row):
                        invoice["due_date"] = row[due_idx]
                    
                    if vat_idx is not None and vat_idx < len(row):
                        invoice["vat"] = float(row[vat_idx] or 0)
                    
                    if status_idx is not None and status_idx < len(row):
                        invoice["payment_status"] = "paid" if row[status_idx] == "Ödenmiş" else "unpaid"
                    
                    # Faturayı ekle
                    self.ledger.add_invoice(invoice)
                    success_count += 1
                    
                except Exception as e:
                    logger.warning(f"Satır içe aktarımında hata: {e}")
                    continue
            
            return True, success_count, ""
            
        except Exception as e:
            logger.error(f"Fatura içe aktarımında hata: {e}")
            return False, 0, str(e)
    
    def _import_chart_of_accounts(self, headers, rows):
        """Hesap planı verilerini içe aktar
        
        Args:
            headers: Başlık satırı
            rows: Veri satırları
            
        Returns:
            tuple: (başarılı mı, içe aktarılan kayıt sayısı, hata mesajı)
        """
        try:
            # Gerekli sütunları doğrula
            required_columns = ["Hesap Kodu", "Hesap Adı", "Tür"]
            for col in required_columns:
                if col not in headers:
                    return False, 0, f"Gerekli sütun bulunamadı: {col}"
            
            # Sütun indekslerini bul
            code_idx = headers.index("Hesap Kodu")
            name_idx = headers.index("Hesap Adı")
            type_idx = headers.index("Tür")
            category_idx = headers.index("Kategori") if "Kategori" in headers else None
            vat_idx = headers.index("KDV Oranı") if "KDV Oranı" in headers else None
            
            # Hesap türü dönüşüm tablosu
            account_type_map = {
                "Varlık": "asset",
                "Borç": "liability",
                "Öz Sermaye": "equity",
                "Gelir": "income",
                "Gider": "expense"
            }
            
            # Kategori dönüşüm tablosu
            category_map = {
                "Dönen Varlık": "current_asset",
                "Duran Varlık": "fixed_asset",
                "Kısa Vadeli Borç": "current_liability",
                "Uzun Vadeli Borç": "long_term_liability",
                "Öz Sermaye": "equity",
                "Hasılat": "revenue",
                "Satış Maliyeti": "cost_of_sales",
                "Faaliyet Gideri": "operating_expense",
                "Finansman Gideri": "financial_expense",
                "Diğer Gelir": "other_income"
            }
            
            # Başarılı içe aktarım sayacı
            success_count = 0
            
            # Her satır için
            for row in rows:
                try:
                    # Boş satırları atla
                    if not row or len(row) < max(code_idx, name_idx, type_idx) + 1:
                        continue
                    
                    # Hesap türünü belirle
                    account_type = account_type_map.get(row[type_idx], "other")
                    
                    # Hesap verilerini oluştur
                    account = {
                        "code": row[code_idx],
                        "name": row[name_idx],
                        "type": account_type
                    }
                    
                    # Opsiyonel alanlar
                    if category_idx is not None and category_idx < len(row):
                        account["category"] = category_map.get(row[category_idx], "other")
                    
                    if vat_idx is not None and vat_idx < len(row):
                        # KDV oranını sayıya çevir
                        vat_str = row[vat_idx]
                        if vat_str:
                            # % işaretini kaldır
                            vat_str = vat_str.replace("%", "").strip()
                            try:
                                account["vat_rate"] = float(vat_str)
                            except ValueError:
                                account["vat_rate"] = 0
                    
                    # Hesabı ekle
                    try:
                        self.ledger.add_account(account)
                        success_count += 1
                    except ValueError as e:
                        # Hesap zaten varsa, güncelle
                        if "zaten kullanılıyor" in str(e):
                            self.ledger.update_account(account["code"], account)
                            success_count += 1
                        else:
                            raise
                    
                except Exception as e:
                    logger.warning(f"Satır içe aktarımında hata: {e}")
                    continue
            
            return True, success_count, ""
            
        except Exception as e:
            logger.error(f"Hesap planı içe aktarımında hata: {e}")
            return False, 0, str(e)
